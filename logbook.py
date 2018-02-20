#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fahrtenbuch zum berechnen der Entfernung zweier Adressen
"""
from time import sleep
from openpyxl import load_workbook
from googlemaps import Client

def calc_distance(destination, target):
    gmaps = Client(key='SomeSecretKey')
    route = gmaps.directions(destination, target)
    legs = route[0]['legs']
    distance = legs[0]['distance']
    return distance['value'] / 1000.0

def main(workbook='Fahrtenbuch.xlsx', sheet='Tabelle1',
         destination_col='A', target_col='B',
         distance_col='C', start_row=2):
    wb = load_workbook(workbook)
    ws = wb.get_sheet_by_name(sheet)
    row = start_row
    destination = ws['{0}{1}'.format(destination_col, row)].value
    while destination:
        target = ws['{0}{1}'.format(target_col, row)].value
        if not target:
            continue
        distance = calc_distance(destination, target)
        ws['{0}{1}'.format(distance_col, row)] = distance
        row += 1
        destination = ws['{0}{1}'.format(destination_col, row)].value
        sleep(5)
    wb.save(workbook)


if __name__ == '__main__':
    main()

